import mimetypes
import requests
import json
from requests.auth import HTTPBasicAuth
from openpyxl import load_workbook
from collections import defaultdict
from lxml import etree
import time


filename = "C://PjtExport.xlsx"
wb = load_workbook(filename)
sheet = wb["Sheet1"]
confdict = defaultdict(list)

for i1 in range(2,100):
    if(str(sheet.cell(row=i1, column=1).value) != "None") :
        spaceName = str(sheet.cell(row=i1, column=1).value)
        keyName = spaceName.replace(" ","")
        #print(keyName)
        url = "https://XXX/wiki/rest/api/space/" + keyName
        auth = HTTPBasicAuth('', '')
        response = requests.request(
            "GET",
            url,
            auth=auth
        )
        # print(response.status_code)
        if response.status_code != 404:
            res_data = response.json()
            homepage = str(res_data['_expandable']['homepage'])
            spaceID = str(homepage[18:])
            pageName = str(sheet.cell(row=i1, column=2).value)
            url = "https://XXX/wiki/rest/api/content?title=" + pageName + "&spaceKey=" + keyName + "/content"
            auth = HTTPBasicAuth('', '')
            response1 = requests.request(
                "GET",
                url,
                auth=auth
            )
            #print(response1.status_code)
            # print(spaceID)
            if response1.status_code == 404:
                #print(i1)
                #print(pageName)

                url = "https://XXX/wiki/rest/api/content/317456545/pagehierarchy/copy"
                auth = HTTPBasicAuth('', '')
                headers = {
                    "Content-Type": "application/json"
                }
                # print(url)
                payload = json.dumps({
                    "copyAttachments": True,
                    "copyPermissions": True,
                    "copyProperties": True,
                    "copyLabels": True,
                    "copyCustomContents": True,
                    "destinationPageId": spaceID,
                    "titleOptions": {
                        "prefix": pageName + "  ",
                        "replace": " ",
                        "search": ""
                    }
                })
                # print(payload)
                response2 = requests.request(
                    "POST",
                    url,
                    data=payload,
                    headers=headers,
                    auth=auth
                )
                # print(response2.text)
        else:
            url = 'https://XXX/wiki/rest/api/space'
            auth = HTTPBasicAuth('', '')
            headers = {
                "Accept": "application/json",
                "Content-Type": "application/json"
            }
            payload = json.dumps({
                "key": keyName,
                "name": spaceName,
                "description": {
                    "plain": {
                        "value": "Test space created in Python - XL Int",
                        "representation": "storage"
                    }
                },
            })
            response = requests.request(
                "POST",
                url,
                data=payload,
                headers=headers,
                auth=auth
            )
            res_data = response.json()
            spaceID = str((res_data["homepage"]["id"]))
            pageName = str(sheet.cell(row=i1, column=2).value)
            # print(spaceID)
            # print(pageName)
            # confdict.setdefault(keyName, spaceName,spaceID)

            url = "https://XXX/wiki/rest/api/content?title=" + pageName + "&spaceKey=" + keyName + "/content"
            auth = HTTPBasicAuth('', '')
            response1 = requests.request(
                "GET",
                url,
                auth=auth
            )
            # print(response1)
            # print(url)
            # print(keyName)

            if response1.status_code == 404:
                url = "https://XXX/wiki/rest/api/content/317456545/pagehierarchy/copy"
                auth = HTTPBasicAuth('', '')
                headers = {
                    "Content-Type": "application/json"
                }
                payload = json.dumps({
                    "copyAttachments": True,
                    "copyPermissions": True,
                    "copyProperties": True,
                    "copyLabels": True,
                    "copyCustomContents": True,
                    "destinationPageId": spaceID,
                    "titleOptions": {
                        "prefix": pageName + "  ",
                        "replace": " ",
                        "search": ""
                    }
                })
                response = requests.request(
                    "POST",
                    url,
                    data=payload,
                    headers=headers,
                    auth=auth
                    )
                time.sleep(6)

spaces = []
for a in range (2,100):
    if(str(sheet.cell(row=a, column=1).value) != "None") :
        spaceName = str(sheet.cell(row=a, column=1).value)
        projectName = str(sheet.cell(a, column=2).value)
        sowFileName = str(sheet.cell(a, column=3).value)
        delFileName = str(sheet.cell(a, column=4).value)
        url = "https://XXX/wiki/rest/api/space/" + spaceName
        auth = HTTPBasicAuth('', '')
        response = requests.request(
            "GET",
            url,
            auth=auth
        )
        res = response.json()
        #print(url)
        #print(res)
        if not(spaces.__contains__(spaceName)):
            spaces.append(spaceName)
            if response.status_code != 404:
                url = "https://XXX/wiki/rest/api/space/" + spaceName + "/content"
                auth = HTTPBasicAuth('', '')
                headers = {"Accept": "application/json"}
                response11 = requests.request("GET", url, auth=auth)
                if response11.status_code != 404:
                    res11 = response11.json()
                    x = 0
                    #print(spaceName)
                    subName = ""
                    #print(len(res11['page']['results']))
                    uploadType = ""
                    sowcntr = 0
                    delcntr = 0
                    while x < len(res11['page']['results']):
                        name = str(res11['page']['results'][x]['title'])
                        print("name - " + name)
                        #print(cntr)
                        if(name.__contains__("SOW")):
                            subName = name[:len(name) - 5]
                            print("subname - " + subName )
                            print("project name - " + projectName)
                            if(subName == projectName ):
                                id = str(res11['page']['results'][x]['id'])
                                filename = sowFileName
                                    #str(sheet.cell(a, column=3).value)
                                uploadType = "SOW"
                        elif(name.__contains__("Deliverable")):
                            subName = name[:len(name) - 12]
                            if(subName == projectName ):
                                id = str(res11['page']['results'][x]['id'])
                                filename = delFileName
                                    #projectName = str(sheet.cell(a, column=4).value)
                                uploadType = "Deliverable"
                                #print(name)
                                #print(filename)
                                #print(id)

                        if (filename != "None" and subName == projectName ):
                            url = 'https://XXX/wiki/rest/api/content/' + id + '/child/attachment/'
                            auth = HTTPBasicAuth('', '')
                            headers = {'X-Atlassian-Token': 'no-check'}
                            if(uploadType == "SOW" and sowcntr == 0 ):
                                file1 = "C:/Users/AshwinKumar/Files/" + spaceName + "/" + projectName + "/SOW/" + filename
                                sowcntr = sowcntr + 1
                                content_type, encoding = mimetypes.guess_type(file1)
                                if content_type is None:
                                    content_type = 'multipart/form-data'
                                files1 = {'file': (file1, open(file1, 'rb'), content_type)}
                                r = requests.post(url, headers=headers, files=files1, auth=auth)
                                print(spaceName + "|" + projectName + "|" + uploadType + "|" + filename + "|" + id)
                            elif(uploadType == "Deliverable" and delcntr == 0):
                                file1 = "C:/Users/AshwinKumar/Files/" + spaceName + "/" + projectName + "/Deliverable/" + filename
                                delcntr = delcntr + 1
                                content_type, encoding = mimetypes.guess_type(file1)
                                if content_type is None:
                                    content_type = 'multipart/form-data'
                                files1 = {'file': (file1, open(file1, 'rb'), content_type)}
                                r = requests.post(url, headers=headers, files=files1, auth=auth)
                                print(spaceName + "|" + projectName + "|" + uploadType + "|" + filename + "|" + id)

                        name = ""

                        x = x + 1
                time.sleep(6)








