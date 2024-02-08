import mimetypes
import requests
import json
from requests.auth import HTTPBasicAuth
from openpyxl import load_workbook
from collections import defaultdict
from lxml import etree
import time


filename = "C:/XXX/PjtExport.xlsx"
wb = load_workbook(filename)
sheet = wb["Sheet1"]
confdict = defaultdict(list)

for i1 in range(2,100):
    if(str(sheet.cell(row=i1, column=1).value) != "None") :
        spaceName = str(sheet.cell(row=i1, column=1).value)
        keyName = spaceName.replace(" ","")
        #print(keyName)
        url = "https://XXX/wiki/rest/api/space/" + keyName
        auth = HTTPBasicAuth('', '')
        response = requests.request(
            "GET",
            url,
            auth=auth
        )
        # print(response.status_code)
        if response.status_code != 404:
            res_data = response.json()
            homepage = str(res_data['_expandable']['homepage'])
            spaceID = str(homepage[18:])
            pageName = str(sheet.cell(row=i1, column=2).value)
            url = "https://XXX/wiki/rest/api/content?title=" + pageName + "&spaceKey=" + keyName + "/content"
            auth = HTTPBasicAuth('', '')
            response1 = requests.request(
                "GET",
                url,
                auth=auth
            )
            #print(response1.status_code)
            # print(spaceID)
            if response1.status_code == 404:
                #print(i1)
                #print(pageName)

                url = "https://XXX/wiki/rest/api/content/XXX/pagehierarchy/copy"
                auth = HTTPBasicAuth('', '')
                headers = {
                    "Content-Type": "application/json"
                }
                # print(url)
                payload = json.dumps({
                    "copyAttachments": True,
                    "copyPermissions": True,
                    "copyProperties": True,
                    "copyLabels": True,
                    "copyCustomContents": True,
                    "destinationPageId": spaceID,
                    "titleOptions": {
                        "prefix": pageName + "  ",
                        "replace": " ",
                        "search": ""
                    }
                })
                # print(payload)
                response2 = requests.request(
                    "POST",
                    url,
                    data=payload,
                    headers=headers,
                    auth=auth
                )
                # print(response2.text)
        else:
            url = 'https://XXX/wiki/rest/api/space'
            auth = HTTPBasicAuth('', '')
            headers = {
                "Accept": "application/json",
                "Content-Type": "application/json"
            }
            payload = json.dumps({
                "key": keyName,
                "name": spaceName,
                "description": {
                    "plain": {
                        "value": "Test space created in Python - XL Int",
                        "representation": "storage"
                    }
                },
            })
            response = requests.request(
                "POST",
                url,
                data=payload,
                headers=headers,
                auth=auth
            )
            res_data = response.json()
            spaceID = str((res_data["homepage"]["id"]))
            pageName = str(sheet.cell(row=i1, column=2).value)
            # print(spaceID)
            # print(pageName)
            # confdict.setdefault(keyName, spaceName,spaceID)

            url = "https://XXX/wiki/rest/api/content?title=" + pageName + "&spaceKey=" + keyName + "/content"
            auth = HTTPBasicAuth('', '')
            response1 = requests.request(
                "GET",
                url,
                auth=auth
            )
            # print(response1)
            # print(url)
            # print(keyName)

            if response1.status_code == 404:
                url = "https://XXX/wiki/rest/api/content/317456545/pagehierarchy/copy"
                auth = HTTPBasicAuth('', '')
                headers = {
                    "Content-Type": "application/json"
                }
                payload = json.dumps({
                    "copyAttachments": True,
                    "copyPermissions": True,
                    "copyProperties": True,
                    "copyLabels": True,
                    "copyCustomContents": True,
                    "destinationPageId": spaceID,
                    "titleOptions": {
                        "prefix": pageName + "  ",
                        "replace": " ",
                        "search": ""
                    }
                })
                response = requests.request(
                    "POST",
                    url,
                    data=payload,
                    headers=headers,
                    auth=auth
                    )
                time.sleep(6)

